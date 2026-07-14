from __future__ import annotations

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ── Paleta ────────────────────────────────────────────────────────────────────

_GRIS_OSCURO      = "2C2C2C"
_GRIS_MEDIO       = "6B6B6B"
_GRIS_CLARO       = "E0E0E0"
_GRIS_BORDE       = "D0D0D0"
_GRIS_FILA_ALTERNA = "F5F5F5"
_BLANCO           = "FFFFFF"
_ACENTO           = "2E6DA4"
_NEGRO            = "000000"

# ── Objetos de estilo reutilizables ───────────────────────────────────────────

_borde_inferior = Border(
    bottom=Side(style="thin", color=_GRIS_BORDE)
)

_borde_encabezado = Border(
    bottom=Side(style="medium", color=_GRIS_OSCURO)
)

_borde_titulo_seccion = Border(
    bottom=Side(style="thin", color=_GRIS_BORDE)
)

_borde_fila_dato = Border(
    left=Side(style="thin", color=_GRIS_BORDE),
    right=Side(style="thin", color=_GRIS_BORDE),
    bottom=Side(style="thin", color=_GRIS_BORDE),
)

# ── Funciones públicas ────────────────────────────────────────────────────────

def aplicar_estilo_titulo_programa(celda) -> None:
    celda.font = Font(name="Arial", size=13, bold=True, color=_GRIS_OSCURO)


def aplicar_estilo_label_metadata(celda) -> None:
    celda.font = Font(name="Arial", size=10, color=_GRIS_MEDIO)


def aplicar_estilo_valor_metadata(celda) -> None:
    celda.font = Font(name="Arial", size=10, bold=False, color=_GRIS_OSCURO)


def aplicar_estilo_titulo_seccion(
    hoja, fila: int, columna: int, num_columnas: int
) -> None:
    num_cols = max(num_columnas, 1)
    for desplazamiento in range(num_cols):
        celda = hoja.cell(row=fila, column=columna + desplazamiento)
        celda.fill = PatternFill("solid", fgColor=_GRIS_CLARO)
        celda.border = _borde_titulo_seccion
        celda.font = Font(name="Arial", size=10, bold=True, color=_GRIS_OSCURO)


def aplicar_estilo_encabezado_tabla(celda) -> None:
    celda.fill = PatternFill("solid", fgColor=_GRIS_OSCURO)
    celda.font = Font(name="Arial", size=9, bold=True, color=_BLANCO)
    celda.alignment = Alignment(horizontal="left", vertical="center")
    celda.border = Border(
        left=Side(style="thin", color=_GRIS_BORDE),
        right=Side(style="thin", color=_GRIS_BORDE),
        bottom=Side(style="medium", color=_GRIS_OSCURO),
    )


def aplicar_estilo_fila_dato(celda, alterna: bool = False) -> None:
    celda.font = Font(name="Arial", size=9, color=_GRIS_OSCURO)
    celda.border = _borde_fila_dato
    celda.alignment = Alignment(
        horizontal="left", vertical="center", wrap_text=False
    )
    if alterna:
        celda.fill = PatternFill("solid", fgColor=_GRIS_FILA_ALTERNA)


def aplicar_estilo_fila_dato_acento(celda, alterna: bool = False) -> None:
    celda.font = Font(name="Arial", size=9, bold=True, color=_ACENTO)
    celda.border = _borde_fila_dato
    celda.alignment = Alignment(horizontal="left", vertical="center")
    if alterna:
        celda.fill = PatternFill("solid", fgColor=_GRIS_FILA_ALTERNA)


def aplicar_estilo_mensaje_vacio(celda) -> None:
    celda.font = Font(name="Arial", size=9, italic=True, color=_GRIS_MEDIO)


def aplicar_estilo_celda_gris(celda) -> None:
    celda.fill = PatternFill("solid", fgColor=_GRIS_CLARO)


_FACTOR_ANCHO_TITULO = 1.4
_MARGEN_ANCHO_TITULO = 2


def ajustar_ancho_columna_titulo(hoja, columna: int, texto: str) -> None:
    """
    Ajusta el ancho de la columna del título del programa (ej. "COMBOS
    v1.0.0") para que el texto no quede recortado.
    """
    ancho_requerido = (
        round(len(texto) * _FACTOR_ANCHO_TITULO) + _MARGEN_ANCHO_TITULO
    )
    hoja.column_dimensions[get_column_letter(columna)].width = ancho_requerido


def ajustar_anchos_columnas(hoja, columna_inicio: int) -> None:
    ancho_minimo = 10
    ancho_maximo = 60

    for columna in hoja.columns:
        ancho = ancho_minimo
        for celda in columna:
            if celda.value is not None:
                ancho = max(ancho, len(str(celda.value)) + 4)
        letra = get_column_letter(columna[0].column)
        # Las columnas anteriores a columna_inicio se dejan angostas (margen).
        if columna[0].column < columna_inicio:
            hoja.column_dimensions[letra].width = 3
        else:
            ancho_existente = hoja.column_dimensions[letra].width or 0
            hoja.column_dimensions[letra].width = min(
                max(ancho, ancho_existente), ancho_maximo
            )


def aplicar_estilo_etiqueta_hoja(celda) -> None:
    celda.font = Font(name="Arial", size=9, bold=True, color=_ACENTO)


def aplicar_borde_perimetral_tabla(
    hoja, fila_inicio: int, fila_fin: int, col_inicio: int, col_fin: int
) -> None:
    borde_negro = Side(style="thin", color=_NEGRO)
    for numero_fila in range(fila_inicio, fila_fin + 1):
        for numero_col in range(col_inicio, col_fin + 1):
            celda = hoja.cell(row=numero_fila, column=numero_col)
            borde_actual = celda.border
            celda.border = Border(
                left=(
                    borde_negro if numero_col == col_inicio
                    else borde_actual.left
                ),
                right=(
                    borde_negro if numero_col == col_fin
                    else borde_actual.right
                ),
                top=(
                    borde_negro if numero_fila == fila_inicio
                    else borde_actual.top
                ),
                bottom=(
                    borde_negro if numero_fila == fila_fin
                    else borde_actual.bottom
                ),
            )
