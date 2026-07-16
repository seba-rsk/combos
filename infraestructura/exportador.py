from __future__ import annotations

from collections import defaultdict
from pathlib import Path

import openpyxl

from dominio.envolventes import construir_envolventes
from dominio.formateador import formatear_componentes
from dominio.modelos import (
    Combinacion,
    Componente,
    EleccionParametro,
    EstadoCrudo,
)
from dominio.sesion import Sesion
from infraestructura.encabezado_excel import escribir_encabezado_programa
from infraestructura.estilos_excel import (
    ajustar_anchos_columnas,
    aplicar_borde_perimetral_tabla,
    aplicar_estilo_encabezado_tabla,
    aplicar_estilo_fila_dato,
    aplicar_estilo_fila_dato_acento,
    aplicar_estilo_mensaje_vacio,
    aplicar_estilo_titulo_seccion,
)
from infraestructura.guardado_excel import guardar_excel_atomico
from infraestructura.sanitizacion_excel import neutralizar_texto_libre

# Límite de Excel para el nombre de una hoja.
MAX_CARACTERES_NOMBRE_HOJA = 31

# ── Punto de entrada ──────────────────────────────────────────────────────────

def exportar(
    sesion: Sesion,
    config_resumen: dict,
    config_exportador: dict,
    ruta_destino: str,
    version: str,
) -> None:
    """
    Genera el archivo Excel de salida con la hoja de resumen del proceso
    completo y la hoja de exportación en el formato configurado.

    Args:
        sesion: Sesión ya procesada y con la decisión de descartes
            aplicada: aporta las combinaciones, los estados leídos, el
            reglamento resuelto, las elecciones de parámetros y el
            nombre del perfil para el encabezado.
        config_resumen: Configuración de secciones y formato de la hoja
            de resumen (ver infraestructura.config_interna.CONFIG_RESUMEN).
        config_exportador: Configuración del perfil de exportación elegido
            (YAML de la carpeta exportadores/).
        ruta_destino: Ruta donde se guarda el archivo Excel generado.
        version: Versión de COMBOS, para mostrarla en el encabezado.

    Raises:
        ValueError: Si config_resumen o config_exportador no tienen los
            campos obligatorios, o si la ruta de destino no es válida.
        RuntimeError: Si el archivo no se pudo guardar en la ruta indicada.
    """
    _validar_config_resumen(config_resumen)
    _validar_config_exportador(config_exportador)
    _validar_ruta_destino(ruta_destino)

    _asignar_nombres(sesion.combinaciones, sesion.reglamento)

    libro = openpyxl.Workbook()
    libro.remove(libro.worksheets[0])

    _escribir_hoja_resumen(
        libro, sesion.combinaciones, sesion.estados_crudos,
        sesion.reglamento, sesion.elecciones,
        config_resumen, sesion.nombre_perfil, version,
    )
    _escribir_hoja_exportacion(
        libro, sesion.combinaciones, config_exportador, sesion.reglamento
    )

    if libro.worksheets:
        libro.active = 0
    try:
        guardar_excel_atomico(libro, ruta_destino)
    except (OSError, PermissionError) as error:
        raise RuntimeError(
            f"No se pudo guardar el archivo en '{ruta_destino}': {error}"
        ) from error


# ── Paso 1: asignación de nombres ─────────────────────────────────────────────

def _asignar_nombres(
    combinaciones: list[Combinacion], reglamento: dict
) -> None:
    combinaciones_validas = [
        c for c in combinaciones
        if not c.es_duplicada and not c.descartada_por_usuario
    ]

    por_estado_limite: dict[str, list[Combinacion]] = defaultdict(list)
    for combinacion in combinaciones_validas:
        por_estado_limite[combinacion.estado_limite].append(combinacion)

    for estado_limite, grupo in por_estado_limite.items():
        prefijo = _obtener_prefijo(reglamento, estado_limite)
        _asignar_nombres_en_grupo(grupo, prefijo)


def _asignar_nombres_en_grupo(
    grupo: list[Combinacion], prefijo: str
) -> None:
    por_base: dict[int, list[Combinacion]] = defaultdict(list)
    for combinacion in grupo:
        por_base[combinacion.combinacion_base_id].append(combinacion)

    numero_secuencial = 1
    for variantes in por_base.values():
        if len(variantes) == 1:
            variantes[0].nombre = f"{prefijo}{numero_secuencial}"
        else:
            for indice_variante, variante in enumerate(variantes, start=1):
                variante.nombre = (
                    f"{prefijo}{numero_secuencial}-{indice_variante}"
                )
        numero_secuencial += 1


def _obtener_prefijo(reglamento: dict, estado_limite: str) -> str:
    config_estado = reglamento.get("limit_states", {}).get(estado_limite, {})
    if isinstance(config_estado, dict):
        return config_estado.get("prefix", estado_limite)
    return estado_limite


# ── Hoja 1: Resumen ───────────────────────────────────────────────────────────

def _escribir_hoja_resumen(
    libro: openpyxl.Workbook,
    combinaciones: list[Combinacion],
    estados_crudos: list[EstadoCrudo],
    reglamento: dict,
    elecciones: list[EleccionParametro],
    config_resumen: dict,
    nombre_perfil: str,
    version: str,
) -> None:
    hoja = libro.create_sheet(config_resumen["nombre_hoja"])
    fila = config_resumen["fila_inicio"]
    columna = config_resumen["columna_inicio"]

    encabezado = config_resumen.get("encabezado_programa", {})
    fila = escribir_encabezado_programa(
        hoja, fila, columna, reglamento, nombre_perfil, version,
        etiqueta="SALIDA DE DATOS",
        nombre_programa=encabezado.get("nombre", "COMBOS"),
        mostrar_reglamento=encabezado.get("mostrar_reglamento", False),
        mostrar_fecha=encabezado.get("mostrar_fecha", False),
        elecciones=elecciones,
    )
    fila += 1

    escritores_por_id = {
        "resumen_ejecutivo": _escribir_seccion_resumen_ejecutivo,
        "datos_ingresados": _escribir_seccion_datos_ingresados,
        "combinaciones_generadas": _escribir_seccion_combinaciones_generadas,
        "combinaciones_resultantes": (
            _escribir_seccion_combinaciones_resultantes
        ),
        "duplicados_eliminados": _escribir_seccion_duplicados_eliminados,
        "superadas": _escribir_seccion_superadas,
    }

    for config_seccion in config_resumen["secciones"]:
        if not config_seccion.get("visible", True):
            continue
        escritor = escritores_por_id.get(config_seccion["id"])
        if escritor is None:
            continue
        fila = escritor(
            hoja, fila, columna, combinaciones, estados_crudos, config_seccion
        )
        fila += 1

    ajustar_anchos_columnas(hoja, columna)


# ── Secciones del resumen ─────────────────────────────────────────────────────

def _escribir_seccion_resumen_ejecutivo(
    hoja, fila, columna, combinaciones, estados, config_seccion
):
    """
    Tabla compacta con los totales por estado límite (generadas,
    duplicadas, superadas descartadas, resultantes), para que alcance
    con leer esta sección sin recorrer el detalle de las demás.
    """
    aplicar_estilo_titulo_seccion(
        hoja, fila, columna, len(config_seccion.get("columnas", []))
    )
    hoja.cell(row=fila, column=columna, value=config_seccion["titulo"])
    fila += 1
    encabezados = config_seccion.get("columnas", [])
    for d, t in enumerate(encabezados):
        celda = hoja.cell(row=fila, column=columna + d, value=t)
        aplicar_estilo_encabezado_tabla(celda)
    fila_encabezado = fila
    fila += 1

    conteos_por_estado = _contar_por_estado_limite(combinaciones)
    indice_fila = 0
    for estado_limite, conteo in conteos_por_estado.items():
        valores = {
            "Estado límite": neutralizar_texto_libre(estado_limite),
            **conteo,
        }
        for d, t in enumerate(encabezados):
            celda = hoja.cell(
                row=fila, column=columna + d, value=valores.get(t, "")
            )
            aplicar_estilo_fila_dato(celda, alterna=indice_fila % 2 == 1)
        fila += 1
        indice_fila += 1

    valores_total = {
        "Estado límite": "TOTAL", **_sumar_conteos(conteos_por_estado)
    }
    for d, t in enumerate(encabezados):
        celda = hoja.cell(
            row=fila, column=columna + d, value=valores_total.get(t, "")
        )
        aplicar_estilo_fila_dato_acento(celda, alterna=indice_fila % 2 == 1)
    fila += 1

    aplicar_borde_perimetral_tabla(
        hoja, fila_encabezado, fila - 1, columna, columna + len(encabezados) - 1
    )
    return fila


def _contar_por_estado_limite(
    combinaciones: list[Combinacion],
) -> dict[str, dict]:
    conteos: dict[str, dict] = {}
    for combinacion in combinaciones:
        conteo = conteos.setdefault(
            combinacion.estado_limite,
            {
                "Generadas": 0,
                "Duplicadas": 0,
                "Superadas (descartadas)": 0,
                "Resultantes": 0,
            },
        )
        es_duplicada = combinacion.es_duplicada
        descartada = combinacion.descartada_por_usuario

        conteo["Generadas"] += 1
        if es_duplicada:
            conteo["Duplicadas"] += 1
        if combinacion.esta_superada and descartada:
            conteo["Superadas (descartadas)"] += 1
        if not es_duplicada and not descartada:
            conteo["Resultantes"] += 1
    return conteos


def _sumar_conteos(conteos_por_estado: dict[str, dict]) -> dict[str, int]:
    total = {
        "Generadas": 0,
        "Duplicadas": 0,
        "Superadas (descartadas)": 0,
        "Resultantes": 0,
    }
    for conteo in conteos_por_estado.values():
        for clave, valor in conteo.items():
            total[clave] += valor
    return total


def _escribir_tabla_seccion(
    hoja,
    fila,
    columna,
    config_seccion,
    registros,
    construir_valores,
    mensaje_vacio=None,
    estilo_columna=None,
):
    """
    Escribe una sección estándar del resumen: título, encabezados y una
    fila por cada registro (con bandas alternas y borde perimetral).
    Las cinco secciones de detalle comparten esta misma estructura y
    solo difieren en qué registros muestran y cómo arman los valores.

    Args:
        registros: Lista ya filtrada de elementos a mostrar, uno por fila.
        construir_valores: Función (indice, registro) -> dict de valores
            de esa fila, indexados por título de columna.
        mensaje_vacio: Si se indica, se muestra en vez de la tabla cuando
            "registros" está vacía.
        estilo_columna: Función opcional (titulo_columna) -> función de
            estilo para esa columna, en vez de aplicar_estilo_fila_dato.
    """
    aplicar_estilo_titulo_seccion(
        hoja, fila, columna, len(config_seccion.get("columnas", []))
    )
    hoja.cell(row=fila, column=columna, value=config_seccion["titulo"])
    fila += 1

    if mensaje_vacio is not None and not registros:
        celda = hoja.cell(row=fila, column=columna, value=mensaje_vacio)
        aplicar_estilo_mensaje_vacio(celda)
        return fila + 1

    encabezados = config_seccion.get("columnas", [])
    for d, t in enumerate(encabezados):
        celda = hoja.cell(row=fila, column=columna + d, value=t)
        aplicar_estilo_encabezado_tabla(celda)
    fila_encabezado = fila
    fila += 1

    for indice, registro in enumerate(registros):
        valores = construir_valores(indice, registro)
        alterna = indice % 2 == 1
        for d, t in enumerate(encabezados):
            celda = hoja.cell(
                row=fila, column=columna + d, value=valores.get(t, "")
            )
            estilo = estilo_columna(t) if estilo_columna else None
            (estilo or aplicar_estilo_fila_dato)(celda, alterna=alterna)
        fila += 1

    aplicar_borde_perimetral_tabla(
        hoja, fila_encabezado, fila - 1, columna, columna + len(encabezados) - 1
    )
    return fila


def _escribir_seccion_datos_ingresados(
    hoja, fila, columna, combinaciones, estados, config_seccion
):
    def construir_valores(indice, estado: EstadoCrudo):
        es_direccional = estado.tipo_estado == "direccional"
        nombre_grupo = (
            f"{estado.tipo_carga}-{estado.grupo}"
            if es_direccional else ""
        )
        opuesto = (
            "" if not es_direccional
            else ("Sí" if estado.incluir_opuesto else "No")
        )
        return {
            "Id": indice + 1,
            "Nombre del estado": neutralizar_texto_libre(estado.nombre_estado),
            "Tipo de carga": neutralizar_texto_libre(estado.tipo_carga),
            "Tipo de estado": estado.tipo_estado.capitalize(),
            "Nombre de grupo": neutralizar_texto_libre(nombre_grupo),
            "Opuesto": opuesto,
        }

    return _escribir_tabla_seccion(
        hoja, fila, columna, config_seccion, estados, construir_valores
    )


def _escribir_seccion_combinaciones_generadas(
    hoja, fila, columna, combinaciones, estados, config_seccion
):
    def construir_valores(indice, combinacion: Combinacion):
        return {
            "Id": combinacion.indice_generacion,
            "Componentes": neutralizar_texto_libre(
                formatear_componentes(combinacion.componentes)
            ),
            "Estado límite": neutralizar_texto_libre(
                combinacion.estado_limite
            ),
            "Combinación base": combinacion.combinacion_base_id,
            "Duplicada por": (
                combinacion.duplicada_por
                if combinacion.es_duplicada else ""
            ),
            "Superada por": (
                combinacion.superada_por
                if combinacion.esta_superada else ""
            ),
            "Decisión": (
                "Descartada"
                if combinacion.esta_superada
                and combinacion.descartada_por_usuario
                else "No descartada" if combinacion.esta_superada else ""
            ),
        }

    return _escribir_tabla_seccion(
        hoja, fila, columna, config_seccion, combinaciones, construir_valores
    )


def _escribir_seccion_combinaciones_resultantes(
    hoja, fila, columna, combinaciones, estados, config_seccion
):
    resultantes = [
        c for c in combinaciones
        if not c.es_duplicada and not c.descartada_por_usuario
    ]

    def construir_valores(indice, combinacion: Combinacion):
        return {
            "Id": combinacion.indice_generacion,
            "Componentes": neutralizar_texto_libre(
                formatear_componentes(combinacion.componentes)
            ),
            "Estado límite": neutralizar_texto_libre(
                combinacion.estado_limite
            ),
            "Combinación base": combinacion.combinacion_base_id,
            "Nombre asignado": neutralizar_texto_libre(
                combinacion.nombre or ""
            ),
        }

    def estilo_columna(titulo):
        if titulo == "Nombre asignado":
            return aplicar_estilo_fila_dato_acento
        return None

    return _escribir_tabla_seccion(
        hoja, fila, columna, config_seccion, resultantes, construir_valores,
        mensaje_vacio="(Sin combinaciones resultantes)",
        estilo_columna=estilo_columna,
    )


def _escribir_seccion_duplicados_eliminados(
    hoja, fila, columna, combinaciones, estados, config_seccion
):
    duplicadas = [c for c in combinaciones if c.es_duplicada]

    def construir_valores(indice, combinacion: Combinacion):
        return {
            "Id": combinacion.indice_generacion,
            "Componentes": neutralizar_texto_libre(
                formatear_componentes(combinacion.componentes)
            ),
            "Estado límite": neutralizar_texto_libre(
                combinacion.estado_limite
            ),
            "Combinación base": combinacion.combinacion_base_id,
            "Duplicada por": combinacion.duplicada_por,
        }

    return _escribir_tabla_seccion(
        hoja, fila, columna, config_seccion, duplicadas, construir_valores,
        mensaje_vacio="(Sin combinaciones duplicadas)",
    )


def _escribir_seccion_superadas(
    hoja, fila, columna, combinaciones, estados, config_seccion
):
    superadas = [c for c in combinaciones if c.esta_superada]
    if config_seccion.get("filtro") == "solo_descartadas":
        superadas = [c for c in superadas if c.descartada_por_usuario]

    def construir_valores(indice, combinacion: Combinacion):
        return {
            "Id": combinacion.indice_generacion,
            "Componentes": neutralizar_texto_libre(
                formatear_componentes(combinacion.componentes)
            ),
            "Estado límite": neutralizar_texto_libre(
                combinacion.estado_limite
            ),
            "Combinación base": combinacion.combinacion_base_id,
            "Superada por": combinacion.superada_por,
        }

    return _escribir_tabla_seccion(
        hoja, fila, columna, config_seccion, superadas, construir_valores,
        mensaje_vacio="(Sin combinaciones superadas)",
    )


# ── Hoja 2: Exportación ───────────────────────────────────────────────────────

def _escribir_hoja_exportacion(
    libro: openpyxl.Workbook,
    combinaciones: list[Combinacion],
    config_exportador: dict,
    reglamento: dict,
) -> None:
    nombre_software = config_exportador["metadata"]["software_name"]
    nombre_hoja_raw = (
        f"Output {nombre_software}".replace("'", "").replace('"', "")
    )
    nombre_hoja = nombre_hoja_raw[:MAX_CARACTERES_NOMBRE_HOJA]
    hoja = libro.create_sheet(nombre_hoja)

    config_hoja = config_exportador["hoja"]
    config_tabla_combinaciones = config_hoja["tabla_combinaciones"]
    layout = config_tabla_combinaciones["layout"]
    fila_inicio = config_tabla_combinaciones.get("fila_inicio", 1)
    columna_inicio = config_tabla_combinaciones.get("columna_inicio", 1)

    combinaciones_validas = [
        c for c in combinaciones
        if not c.es_duplicada and not c.descartada_por_usuario
    ]

    if layout == "por_componente":
        ancho_tabla_1 = _escribir_layout_por_componente(
            hoja, combinaciones_validas, config_tabla_combinaciones,
            fila_inicio, columna_inicio,
        )
    elif layout == "por_combinacion":
        ancho_tabla_1 = _escribir_layout_por_combinacion(
            hoja, combinaciones_validas, config_tabla_combinaciones,
            fila_inicio, columna_inicio,
        )
    else:
        ancho_tabla_1 = 0

    columna_siguiente = columna_inicio + ancho_tabla_1

    config_tabla_nombres = config_hoja.get("tabla_nombres")
    if config_tabla_nombres:
        separacion = config_tabla_nombres.get("separacion_columnas", 1)
        columna_tabla_nombres = columna_siguiente + separacion
        _escribir_tabla_nombres(
            hoja, combinaciones_validas, config_tabla_nombres,
            fila_inicio, columna_tabla_nombres,
        )
        columna_siguiente = columna_tabla_nombres + 1

    config_tabla_envolventes = config_hoja.get("tabla_envolventes")
    if config_tabla_envolventes:
        separacion = config_tabla_envolventes.get("separacion_columnas", 1)
        columna_tabla_envolventes = columna_siguiente + separacion
        filas_envolventes = construir_envolventes(
            combinaciones_validas,
            reglamento,
            config_tabla_envolventes.get("prefijo_nombre", "ENV"),
        )
        _escribir_tabla_envolventes(
            hoja, filas_envolventes, config_tabla_envolventes,
            fila_inicio, columna_tabla_envolventes,
        )

    ajustar_anchos_columnas(hoja, columna_inicio)


def _neutralizar_si_texto(valor):
    """
    Aplica neutralizar_texto_libre solo si el valor es texto. Para los
    títulos definidos en el YAML del exportador, que son texto libre de
    terceros pero podrían venir como número.
    """
    return neutralizar_texto_libre(valor) if isinstance(valor, str) else valor


# ── Layout por_componente ─────────────────────────────────────────────────────

def _escribir_layout_por_componente(
    hoja, combinaciones_validas, config_hoja, fila_inicio, columna_inicio
) -> int:
    columnas = config_hoja["columnas"]
    for desplazamiento, columna in enumerate(columnas):
        celda = hoja.cell(
            row=fila_inicio,
            column=columna_inicio + desplazamiento,
            value=_neutralizar_si_texto(columna["titulo"]),
        )
        aplicar_estilo_encabezado_tabla(celda)
    fila = fila_inicio + 1
    indice_fila = 0
    for combinacion in combinaciones_validas:
        for componente in combinacion.componentes:
            for desplazamiento, columna in enumerate(columnas):
                valor = _resolver_fuente(
                    columna["fuente"], combinacion, componente
                )
                celda = hoja.cell(
                    row=fila,
                    column=columna_inicio + desplazamiento,
                    value=valor,
                )
                aplicar_estilo_fila_dato(celda, alterna=indice_fila % 2 == 1)
            fila += 1
            indice_fila += 1
    return len(columnas)


# ── Layout por_combinacion ────────────────────────────────────────────────────

def _escribir_layout_por_combinacion(
    hoja, combinaciones_validas, config_hoja, fila_inicio, columna_inicio
) -> int:
    titulo_combinacion = config_hoja.get("titulo_combinacion", "Case")

    nombres_estados = _extraer_nombres_estados_en_orden(combinaciones_validas)

    celda = hoja.cell(
        row=fila_inicio,
        column=columna_inicio,
        value=_neutralizar_si_texto(titulo_combinacion),
    )
    aplicar_estilo_encabezado_tabla(celda)
    for desplazamiento, nombre in enumerate(nombres_estados, start=1):
        celda = hoja.cell(
            row=fila_inicio,
            column=columna_inicio + desplazamiento,
            value=neutralizar_texto_libre(nombre),
        )
        aplicar_estilo_encabezado_tabla(celda)

    fila = fila_inicio + 1
    for indice, combinacion in enumerate(combinaciones_validas):
        alterna = indice % 2 == 1
        celda = hoja.cell(
            row=fila,
            column=columna_inicio,
            value=neutralizar_texto_libre(combinacion.nombre),
        )
        aplicar_estilo_fila_dato_acento(celda, alterna=alterna)
        indice_por_estado = {
            c.nombre_estado: c.factor * c.signo
            for c in combinacion.componentes
        }
        for desplazamiento, nombre in enumerate(nombres_estados, start=1):
            valor = indice_por_estado.get(nombre, None)
            celda = hoja.cell(
                row=fila, column=columna_inicio + desplazamiento, value=valor
            )
            aplicar_estilo_fila_dato(celda, alterna=alterna)
        fila += 1

    return 1 + len(nombres_estados)


def _extraer_nombres_estados_en_orden(
    combinaciones_validas: list[Combinacion],
) -> list[str]:
    vistos = set()
    nombres = []
    for combinacion in combinaciones_validas:
        for componente in combinacion.componentes:
            nombre = componente.nombre_estado
            if nombre not in vistos:
                vistos.add(nombre)
                nombres.append(nombre)
    return nombres


def _resolver_fuente(
    fuente: str, combinacion: Combinacion, componente: Componente
):
    if fuente == "nombre_combinacion":
        return neutralizar_texto_libre(combinacion.nombre)
    if fuente == "nombre_estado":
        return neutralizar_texto_libre(componente.nombre_estado)
    if fuente == "factor_por_signo":
        return componente.factor * componente.signo
    raise ValueError(f"Fuente de columna desconocida: '{fuente}'")


# ── Tabla 2: nombres de combinaciones ─────────────────────────────────────────

def _escribir_tabla_nombres(
    hoja,
    combinaciones_validas: list[Combinacion],
    config_tabla: dict,
    fila_inicio: int,
    columna_inicio: int,
) -> None:
    celda = hoja.cell(
        row=fila_inicio,
        column=columna_inicio,
        value=_neutralizar_si_texto(config_tabla["titulo"]),
    )
    aplicar_estilo_encabezado_tabla(celda)
    fila = fila_inicio + 1
    for indice, combinacion in enumerate(combinaciones_validas):
        celda = hoja.cell(
            row=fila,
            column=columna_inicio,
            value=neutralizar_texto_libre(combinacion.nombre),
        )
        aplicar_estilo_fila_dato(celda, alterna=indice % 2 == 1)
        fila += 1


# ── Tabla 3: envolventes ──────────────────────────────────────────────────────

def _escribir_tabla_envolventes(
    hoja,
    filas_envolventes: list,
    config_tabla: dict,
    fila_inicio: int,
    columna_inicio: int,
) -> None:
    columnas = config_tabla["columnas"]
    for desplazamiento, columna in enumerate(columnas):
        celda = hoja.cell(
            row=fila_inicio,
            column=columna_inicio + desplazamiento,
            value=_neutralizar_si_texto(columna["titulo"]),
        )
        aplicar_estilo_encabezado_tabla(celda)

    atributos_por_id = {
        "envelope": "nombre_envolvente",
        "combination": "nombre_combinacion",
        "factor": "factor",
    }

    fila = fila_inicio + 1
    for indice, fila_dato in enumerate(filas_envolventes):
        for desplazamiento, columna in enumerate(columnas):
            atributo = atributos_por_id.get(columna["id"])
            valor = getattr(fila_dato, atributo, None) if atributo else None
            if isinstance(valor, str):
                valor = neutralizar_texto_libre(valor)
            celda = hoja.cell(
                row=fila, column=columna_inicio + desplazamiento, value=valor
            )
            aplicar_estilo_fila_dato(celda, alterna=indice % 2 == 1)
        fila += 1


# ── Validaciones ──────────────────────────────────────────────────────────────

def _validar_config_resumen(config_resumen: dict) -> None:
    campos_requeridos = [
        "nombre_hoja", "fila_inicio", "columna_inicio", "secciones"
    ]
    for campo in campos_requeridos:
        if campo not in config_resumen:
            raise ValueError(
                f"config_resumen: falta el campo obligatorio '{campo}'."
            )
    for seccion in config_resumen["secciones"]:
        if "id" not in seccion or "titulo" not in seccion:
            raise ValueError(
                "config_resumen: cada sección debe tener los campos "
                "'id' y 'titulo'."
            )


def _validar_config_exportador(config_exportador: dict) -> None:
    if not isinstance(config_exportador.get("metadata"), dict):
        raise ValueError(
            "config_exportador: falta el campo 'metadata' o no es una "
            "sección con claves."
        )
    if "software_name" not in config_exportador["metadata"]:
        raise ValueError(
            "config_exportador: falta el campo 'metadata.software_name'."
        )
    if not isinstance(config_exportador.get("hoja"), dict):
        raise ValueError(
            "config_exportador: falta el campo 'hoja' o no es una "
            "sección con claves."
        )
    config_tabla_combinaciones = config_exportador["hoja"].get(
        "tabla_combinaciones"
    )
    if not isinstance(config_tabla_combinaciones, dict):
        raise ValueError(
            "config_exportador: falta el campo 'hoja.tabla_combinaciones' "
            "o no es una sección con claves."
        )
    _validar_layout_tabla_combinaciones(config_tabla_combinaciones)


LAYOUTS_TABLA_COMBINACIONES = ("por_componente", "por_combinacion")


def _validar_layout_tabla_combinaciones(config_tabla: dict) -> None:
    if "layout" not in config_tabla:
        raise ValueError(
            "config_exportador: falta el campo "
            "'hoja.tabla_combinaciones.layout'."
        )
    if config_tabla["layout"] not in LAYOUTS_TABLA_COMBINACIONES:
        raise ValueError(
            f"config_exportador: 'hoja.tabla_combinaciones.layout' tiene "
            f"un valor inválido: '{config_tabla['layout']}'. "
            f"Valores posibles: {LAYOUTS_TABLA_COMBINACIONES}."
        )
    if config_tabla["layout"] == "por_componente":
        _validar_columnas_por_componente(config_tabla.get("columnas"))


def _validar_columnas_por_componente(columnas) -> None:
    if not isinstance(columnas, list):
        raise ValueError(
            "config_exportador: layout 'por_componente' requiere el "
            "campo 'hoja.tabla_combinaciones.columnas' como lista."
        )
    for columna in columnas:
        if not isinstance(columna, dict):
            raise ValueError(
                "config_exportador: cada elemento de "
                "'hoja.tabla_combinaciones.columnas' debe ser una sección "
                "con los campos 'id', 'titulo' y 'fuente'."
            )
        for campo in ("id", "titulo", "fuente"):
            if campo not in columna:
                raise ValueError(
                    f"config_exportador: la columna "
                    f"'{columna.get('id', '?')}' no tiene el campo "
                    f"obligatorio '{campo}'."
                )


def _validar_ruta_destino(ruta_destino: str) -> None:
    ruta = Path(ruta_destino)
    if not ruta.parent.exists():
        raise ValueError(
            f"El directorio de destino no existe: '{ruta.parent}'."
        )
    if ruta.exists() and not ruta.is_file():
        raise ValueError(
            f"La ruta de destino existe pero no es un archivo: "
            f"'{ruta_destino}'."
        )
