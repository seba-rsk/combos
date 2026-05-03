from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.styles.protection import Protection
from openpyxl.worksheet.datavalidation import DataValidation
from infraestructura.estilos_excel import (
    aplicar_estilo_etiqueta_hoja,
    aplicar_estilo_titulo_programa,
    aplicar_estilo_label_metadata,
    aplicar_estilo_valor_metadata,
    aplicar_estilo_encabezado_tabla,
    ajustar_anchos_columnas,
    aplicar_estilo_celda_gris,
    aplicar_borde_perimetral_tabla,
)
from infraestructura.config_interna import CONFIG_PLANTILLA

NOMBRE_HOJA = "Estados de carga"


def generar_plantilla(reglamento: dict, ruta_destino: str, nombre_perfil: str, version: str) -> None:
    """
    Genera el archivo Excel de plantilla para que el usuario complete
    sus estados de carga, basándose en el reglamento y la configuración
    interna de plantilla.

    Raises:
        PermissionError / OSError: Si la ruta de destino no es escribible.
    """
    libro = _construir_libro_excel(CONFIG_PLANTILLA, reglamento, nombre_perfil, version)
    _guardar_archivo(libro, ruta_destino)


def _construir_libro_excel(configuracion: dict, reglamento: dict, nombre_perfil: str, version: str) -> Workbook:
    libro = Workbook()
    hoja = libro.active
    hoja.title = NOMBRE_HOJA

    fila_inicio = configuracion["fila_inicio"]
    fila_tabla = configuracion["fila_tabla"]
    columna_inicio = configuracion["columna_inicio"]
    filas_datos = configuracion["filas_preestablecidas"]
    columnas = configuracion["columnas"]
    fila_inicio_datos = fila_tabla + 1

    _escribir_encabezado_programa(
        hoja, fila_inicio, columna_inicio, reglamento, nombre_perfil, version
    )
    _escribir_encabezado(hoja, columnas, fila_tabla, columna_inicio)
    _escribir_panel_referencia(hoja, fila_tabla, columna_inicio, columnas, reglamento, configuracion["panel_referencia"])

    if configuracion["encabezado"]["borde"]:
        _aplicar_borde_rango(hoja, fila_tabla, fila_tabla, columna_inicio, columna_inicio + len(columnas) - 1)

    _preparar_celdas_desbloqueadas(hoja, fila_tabla, fila_inicio_datos, filas_datos, columna_inicio, columnas)

    for numero_fila in range(fila_inicio_datos, fila_inicio_datos + filas_datos):
        if configuracion["filas"]["borde"]:
            _aplicar_borde_rango(hoja, numero_fila, numero_fila, columna_inicio, columna_inicio + len(columnas) - 1)

    for indice_columna, columna in enumerate(columnas):
        col_excel = columna_inicio + indice_columna
        _configurar_columna(
            hoja, columna, col_excel, fila_inicio_datos, filas_datos, reglamento
        )
    
    aplicar_borde_perimetral_tabla(
        hoja,
        fila_tabla,
        fila_inicio_datos + filas_datos - 1,
        columna_inicio,
        columna_inicio + len(columnas) - 1,
    )
    
    ajustar_anchos_columnas(hoja, columna_inicio)
    _proteger_hoja(hoja)

    return libro


def _escribir_encabezado_programa(
    hoja,
    fila: int,
    columna: int,
    reglamento: dict,
    nombre_perfil: str,
    version: str,
) -> None:
    metadata = reglamento.get("metadata", {})

    celda_titulo = hoja.cell(row=fila, column=columna, value=f"COMBOS v{version}")
    aplicar_estilo_titulo_programa(celda_titulo)
    ancho_requerido = round(len(celda_titulo.value) * 1.4) + 2
    hoja.column_dimensions[get_column_letter(columna)].width = ancho_requerido

    columna_etiqueta = columna + 1
    celda_etiqueta = hoja.cell(row=fila, column=columna_etiqueta, value="ENTRADA DE DATOS")
    aplicar_estilo_etiqueta_hoja(celda_etiqueta)
    fila += 1

    aplicar_estilo_label_metadata(hoja.cell(row=fila, column=columna, value="Perfil usado:"))
    aplicar_estilo_valor_metadata(hoja.cell(row=fila, column=columna + 1, value=nombre_perfil))
    fila += 1

    code_name = metadata.get("code_name", "")
    code_version = metadata.get("code_version", "")
    valor_reglamento = f"{code_name} - {code_version}" if code_name or code_version else ""
    aplicar_estilo_label_metadata(hoja.cell(row=fila, column=columna, value="Reglamento:"))
    aplicar_estilo_valor_metadata(hoja.cell(row=fila, column=columna + 1, value=valor_reglamento))
    fila += 1

    aplicar_estilo_label_metadata(hoja.cell(row=fila, column=columna, value="País:"))
    aplicar_estilo_valor_metadata(hoja.cell(row=fila, column=columna + 1, value=metadata.get("country", "")))
    fila += 1

    aplicar_estilo_label_metadata(hoja.cell(row=fila, column=columna, value="Descripción:"))
    aplicar_estilo_valor_metadata(hoja.cell(row=fila, column=columna + 1, value=metadata.get("description", "")))
    fila += 1

    aplicar_estilo_label_metadata(hoja.cell(row=fila, column=columna, value="Fecha:"))
    aplicar_estilo_valor_metadata(
        hoja.cell(row=fila, column=columna + 1, value=datetime.now().strftime("%d/%m/%Y %H:%M"))
    )


def _escribir_encabezado(
    hoja, columnas: list, fila_encabezado: int, columna_inicio: int
) -> None:
    for indice, columna in enumerate(columnas):
        celda = hoja.cell(row=fila_encabezado, column=columna_inicio + indice)
        celda.value = columna["titulo"]
        aplicar_estilo_encabezado_tabla(celda)
        celda.protection = Protection(locked=True)


def _preparar_celdas_desbloqueadas(
    hoja,
    fila_encabezado: int,
    fila_inicio_datos: int,
    filas_datos: int,
    columna_inicio: int,
    columnas: list,
) -> None:
    """
    Desbloquea todas las celdas de datos para que la protección de hoja
    solo bloquee las celdas autonuméricas, que se re-bloquean luego.
    """
    proteccion_desbloqueada = Protection(locked=False)
    for numero_fila in range(fila_inicio_datos, fila_inicio_datos + filas_datos):
        for indice_columna in range(len(columnas)):
            celda = hoja.cell(row=numero_fila, column=columna_inicio + indice_columna)
            celda.protection = proteccion_desbloqueada


def _configurar_columna(
    hoja,
    columna: dict,
    col_excel: int,
    fila_inicio_datos: int,
    filas_datos: int,
    reglamento: dict,
) -> None:
    tipo = columna["tipo"]

    if tipo == "autonumerico":
        _escribir_autonumerico(hoja, col_excel, fila_inicio_datos, filas_datos)
    elif tipo == "lista_desplegable":
        _aplicar_validacion_lista(hoja, columna, col_excel, fila_inicio_datos, filas_datos, reglamento)
    elif tipo == "entero_positivo":
        _aplicar_validacion_entero_positivo(hoja, columna, col_excel, fila_inicio_datos, filas_datos)

    if columna.get("fondo_gris_fijo"):
        _aplicar_fondo_gris(hoja, col_excel, fila_inicio_datos, filas_datos)


def _escribir_autonumerico(
    hoja, col_excel: int, fila_inicio_datos: int, filas_datos: int
) -> None:
    proteccion_bloqueada = Protection(locked=True)
    estilo_alineacion = Alignment(horizontal="center")

    for numero_fila in range(fila_inicio_datos, fila_inicio_datos + filas_datos):
        numero_secuencial = numero_fila - fila_inicio_datos + 1
        celda = hoja.cell(row=numero_fila, column=col_excel)
        celda.value = numero_secuencial
        celda.protection = proteccion_bloqueada
        celda.alignment = estilo_alineacion


def _aplicar_validacion_lista(
    hoja,
    columna: dict,
    col_excel: int,
    fila_inicio_datos: int,
    filas_datos: int,
    reglamento: dict,
) -> None:
    opciones = _construir_opciones_lista(columna, reglamento)
    formula_opciones = '"' + ",".join(opciones) + '"'
    tipo_error = "stop" if columna.get("validacion_stop") else "warning"
    mensaje_error = columna.get("mensaje_error") or ""

    validacion = DataValidation(
        type="list",
        formula1=formula_opciones,
        allow_blank=True,
        showErrorMessage=bool(mensaje_error),
        error=mensaje_error,
        errorStyle=tipo_error,
    )
    hoja.add_data_validation(validacion)

    rango = _construir_referencia_rango_columna(col_excel, fila_inicio_datos, filas_datos)
    validacion.sqref = rango


def _construir_opciones_lista(columna: dict, reglamento: dict) -> list:
    if columna.get("fuente") == "yaml_load_types":
        return list(reglamento["load_types"].keys())
    return columna.get("opciones", [])


def _aplicar_validacion_entero_positivo(
    hoja,
    columna: dict,
    col_excel: int,
    fila_inicio_datos: int,
    filas_datos: int,
) -> None:
    mensaje_error = columna.get("mensaje_error") or "Ingrese un número entero positivo."

    validacion = DataValidation(
        type="whole",
        operator="greaterThan",
        formula1="0",
        allow_blank=True,
        showErrorMessage=True,
        error=mensaje_error,
        errorStyle="warning",
    )
    hoja.add_data_validation(validacion)

    rango = _construir_referencia_rango_columna(col_excel, fila_inicio_datos, filas_datos)
    validacion.sqref = rango


def _aplicar_fondo_gris(
    hoja, col_excel: int, fila_inicio_datos: int, filas_datos: int
) -> None:
    for numero_fila in range(fila_inicio_datos, fila_inicio_datos + filas_datos):
        aplicar_estilo_celda_gris(hoja.cell(row=numero_fila, column=col_excel))


def _aplicar_borde_rango(
    hoja, fila_inicio: int, fila_fin: int, col_inicio: int, col_fin: int
) -> None:
    borde_fino = Side(style="thin")

    for numero_fila in range(fila_inicio, fila_fin + 1):
        for numero_col in range(col_inicio, col_fin + 1):
            celda = hoja.cell(row=numero_fila, column=numero_col)
            celda.border = Border(
                left=borde_fino,
                right=borde_fino,
                top=borde_fino,
                bottom=borde_fino,
            )


def _construir_referencia_rango_columna(
    col_excel: int, fila_inicio: int, cantidad_filas: int
) -> str:
    letra = get_column_letter(col_excel)
    return f"{letra}{fila_inicio}:{letra}{fila_inicio + cantidad_filas - 1}"


def _proteger_hoja(hoja) -> None:
    hoja.protection.sheet = True
    hoja.protection.selectLockedCells = False
    hoja.protection.selectUnlockedCells = False
    hoja.protection.formatColumns = False


def _escribir_panel_referencia(
    hoja,
    fila_encabezado: int,
    columna_inicio: int,
    columnas: list,
    reglamento: dict,
    configuracion_panel: dict,
) -> None:
    columna_panel = columna_inicio + len(columnas) + configuracion_panel["columnas_separador"]
    fila_actual = fila_encabezado

    celda_titulo_permanentes = hoja.cell(
        row=fila_actual, column=columna_panel, value=configuracion_panel["titulo_permanentes"]
    )
    aplicar_estilo_encabezado_tabla(celda_titulo_permanentes)
    fila_actual += 1

    for nombre_tipo in reglamento["permanent_load_types"]:
        aplicar_estilo_label_metadata(hoja.cell(row=fila_actual, column=columna_panel, value=nombre_tipo))
        fila_actual += 1

    fila_actual += 1

    celda_titulo_tipos = hoja.cell(
        row=fila_actual, column=columna_panel, value=configuracion_panel["titulo_tipos_de_carga"]
    )
    aplicar_estilo_encabezado_tabla(celda_titulo_tipos)
    fila_actual += 1

    for id_tipo, datos_tipo in reglamento["load_types"].items():
        aplicar_estilo_label_metadata(hoja.cell(row=fila_actual, column=columna_panel, value=f"{id_tipo} - {datos_tipo['name']}"))
        fila_actual += 1


def _guardar_archivo(libro: Workbook, ruta_destino: str) -> None:
    from pathlib import Path
    ruta = Path(ruta_destino)
    try:
        libro.save(ruta)
    except (PermissionError, OSError) as error:
        raise type(error)(
            f"No se pudo guardar la plantilla en '{ruta_destino}': {error}"
        ) from error
