import os
import openpyxl


class ErrorArchivoExcel(Exception):
    pass


class ErrorFormatoPlantilla(Exception):
    pass


class ErrorDatoFila(Exception):
    pass


def leer_excel(ruta_excel: str, config_plantilla: dict) -> list[dict]:
    libro = _abrir_libro(ruta_excel)
    hoja = libro.active
    titulos_esperados = _extraer_titulos_esperados(config_plantilla)
    mapa_columnas = _mapear_columnas_por_titulo(hoja, titulos_esperados, config_plantilla)
    fila_datos = config_plantilla["fila_tabla"] + 1
    filas = _leer_filas(hoja, mapa_columnas, fila_datos)
    return filas


def _abrir_libro(ruta_excel: str):
    if not os.path.exists(ruta_excel):
        raise ErrorArchivoExcel(f"No se encontró el archivo: {ruta_excel}")
    try:
        return openpyxl.load_workbook(ruta_excel, data_only=True)
    except Exception as error:
        raise ErrorArchivoExcel(f"No se pudo abrir el archivo Excel: {error}") from error


def _extraer_titulos_esperados(config_plantilla: dict) -> list[str]:
    return [
        columna["titulo"]
        for columna in config_plantilla["columnas"]
        if columna["tipo"] != "autonumerico"
    ]


def _mapear_columnas_por_titulo(
    hoja, titulos_esperados: list[str], config_plantilla: dict
) -> dict[str, int]:
    fila_encabezado = config_plantilla["fila_tabla"]
    columna_inicio = config_plantilla["columna_inicio"]

    titulos_en_hoja = {
        celda.value: celda.column
        for celda in hoja[fila_encabezado]
        if celda.column >= columna_inicio and celda.value is not None
    }

    titulos_faltantes = [t for t in titulos_esperados if t not in titulos_en_hoja]
    if titulos_faltantes:
        raise ErrorFormatoPlantilla(
            f"El archivo no tiene el formato de plantilla COMBOS. "
            f"Columnas faltantes: {', '.join(titulos_faltantes)}"
        )

    return {titulo: titulos_en_hoja[titulo] for titulo in titulos_esperados}


def _leer_filas(hoja, mapa_columnas: dict[str, int], fila_inicio: int) -> list[dict]:
    filas = []
    for numero_fila in range(fila_inicio, hoja.max_row + 1):
        valores = _extraer_valores_fila(hoja, numero_fila, mapa_columnas)
        if _fila_esta_vacia(valores):
            continue
        fila_procesada = _procesar_fila(valores, numero_fila)
        filas.append(fila_procesada)
    return filas


def _extraer_valores_fila(hoja, numero_fila: int, mapa_columnas: dict[str, int]) -> dict:
    return {
        titulo: hoja.cell(row=numero_fila, column=indice_columna).value
        for titulo, indice_columna in mapa_columnas.items()
    }


def _fila_esta_vacia(valores: dict) -> bool:
    return all(valor is None or str(valor).strip() == "" for valor in valores.values())


def _procesar_fila(valores: dict, numero_fila: int) -> dict:
    nombre_estado = _leer_nombre_estado(valores)
    tipo_carga = _leer_tipo_carga(valores)
    tipo_estado = _leer_tipo_estado(valores, numero_fila)

    if tipo_estado == "simple":
        return {
            "nombre_estado": nombre_estado,
            "tipo_carga": tipo_carga,
            "tipo_estado": tipo_estado,
            "grupo": None,
            "incluir_opuesto": False,
        }

    grupo = _leer_grupo_direccional(valores, numero_fila)
    incluir_opuesto = _leer_incluir_opuesto(valores, numero_fila)

    return {
        "nombre_estado": nombre_estado,
        "tipo_carga": tipo_carga,
        "tipo_estado": tipo_estado,
        "grupo": grupo,
        "incluir_opuesto": incluir_opuesto,
    }


def _leer_nombre_estado(valores: dict) -> str:
    return str(valores["Nombre del estado"]).strip() if valores["Nombre del estado"] else ""


def _leer_tipo_carga(valores: dict) -> str:
    return str(valores["Tipo de carga"]).strip() if valores["Tipo de carga"] else ""


def _leer_tipo_estado(valores: dict, numero_fila: int) -> str:
    valor_crudo = valores.get("Tipo de estado")
    if valor_crudo is None or str(valor_crudo).strip() == "":
        raise ErrorDatoFila(
            f"Fila {numero_fila}: el campo 'Tipo de estado' es obligatorio."
        )
    valor_normalizado = str(valor_crudo).strip().lower()
    if valor_normalizado not in ("simple", "direccional"):
        raise ErrorDatoFila(
            f"Fila {numero_fila}: 'Tipo de estado' tiene un valor inválido: '{valor_crudo}'. "
            f"Se esperaba 'Simple' o 'Direccional'."
        )
    return valor_normalizado


def _leer_grupo_direccional(valores: dict, numero_fila: int) -> int:
    valor_crudo = valores.get("Número de grupo")
    if valor_crudo is None or str(valor_crudo).strip() == "":
        raise ErrorDatoFila(
            f"Fila {numero_fila}: estado Direccional requiere un número de grupo."
        )
    return _convertir_a_entero_positivo(valor_crudo, numero_fila)


def _convertir_a_entero_positivo(valor_crudo, numero_fila: int) -> int:
    try:
        valor_numerico = int(float(str(valor_crudo)))
    except (ValueError, TypeError):
        raise ErrorDatoFila(
            f"Fila {numero_fila}: 'Número de grupo' no es un número entero: '{valor_crudo}'."
        )
    if valor_numerico <= 0:
        raise ErrorDatoFila(
            f"Fila {numero_fila}: 'Número de grupo' debe ser un entero positivo, "
            f"pero se recibió: {valor_numerico}."
        )
    return valor_numerico


def _leer_incluir_opuesto(valores: dict, numero_fila: int) -> bool:
    valor_crudo = valores.get("Incluir opuesto")
    if valor_crudo is None or str(valor_crudo).strip() == "":
        raise ErrorDatoFila(
            f"Fila {numero_fila}: estado Direccional requiere un valor en 'Incluir opuesto'."
        )
    valor_normalizado = str(valor_crudo).strip().lower()
    if valor_normalizado not in ("sí", "si", "no"):
        raise ErrorDatoFila(
            f"Fila {numero_fila}: 'Incluir opuesto' tiene un valor inválido: '{valor_crudo}'. "
            f"Se esperaba 'Sí' o 'No'."
        )
    return valor_normalizado in ("sí", "si")
