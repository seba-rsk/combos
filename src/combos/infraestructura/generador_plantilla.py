from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Side
from openpyxl.styles.protection import Protection
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from combos.infraestructura.config_interna import CONFIG_PLANTILLA
from combos.infraestructura.encabezado_excel import escribir_encabezado_programa
from combos.infraestructura.estilos_excel import (
    ajustar_anchos_columnas,
    aplicar_borde_perimetral_tabla,
    aplicar_estilo_celda_gris,
    aplicar_estilo_encabezado_tabla,
    aplicar_estilo_label_metadata,
)
from combos.infraestructura.guardado_excel import guardar_excel_atomico
from combos.infraestructura.sanitizacion_excel import neutralizar_texto_libre

NOMBRE_HOJA = "Estados de carga"


def generar_plantilla(
    reglamento: dict, ruta_destino: str, nombre_perfil: str, version: str
) -> None:
    """
    Genera el archivo Excel de plantilla para que el usuario complete
    sus estados de carga, basándose en el reglamento y la configuración
    interna de plantilla.

    Raises:
        PermissionError / OSError: Si la ruta de destino no es escribible.
    """
    libro = _construir_libro_excel(
        CONFIG_PLANTILLA, reglamento, nombre_perfil, version
    )
    _guardar_archivo(libro, ruta_destino)


def _construir_libro_excel(
    configuracion: dict, reglamento: dict, nombre_perfil: str, version: str
) -> Workbook:
    libro = Workbook()
    hoja = libro.active
    hoja.title = NOMBRE_HOJA

    fila_inicio = configuracion["fila_inicio"]
    fila_tabla = configuracion["fila_tabla"]
    columna_inicio = configuracion["columna_inicio"]
    filas_datos = configuracion["filas_preestablecidas"]
    columnas = configuracion["columnas"]
    fila_inicio_datos = fila_tabla + 1

    escribir_encabezado_programa(
        hoja, fila_inicio, columna_inicio, reglamento, nombre_perfil, version,
        etiqueta="ENTRADA DE DATOS",
    )
    _escribir_encabezado(hoja, columnas, fila_tabla, columna_inicio)
    _escribir_panel_referencia(
        hoja,
        fila_tabla,
        columna_inicio,
        columnas,
        reglamento,
        configuracion["panel_referencia"],
    )

    if configuracion["encabezado"]["borde"]:
        _aplicar_borde_rango(
            hoja, fila_tabla, fila_tabla, columna_inicio,
            columna_inicio + len(columnas) - 1,
        )

    _preparar_celdas_desbloqueadas(
        hoja, fila_tabla, fila_inicio_datos, filas_datos,
        columna_inicio, columnas,
    )

    fila_fin_datos = fila_inicio_datos + filas_datos
    for numero_fila in range(fila_inicio_datos, fila_fin_datos):
        if configuracion["filas"]["borde"]:
            _aplicar_borde_rango(
                hoja, numero_fila, numero_fila, columna_inicio,
                columna_inicio + len(columnas) - 1,
            )

    for indice_columna, columna in enumerate(columnas):
        col_excel = columna_inicio + indice_columna
        _configurar_columna(
            hoja, columna, col_excel, fila_inicio_datos, filas_datos, reglamento
        )

    aplicar_borde_perimetral_tabla(
        hoja,
        fila_tabla,
        fila_inicio_datos + filas_datos - 1,
        columna_inicio,
        columna_inicio + len(columnas) - 1,
    )

    ajustar_anchos_columnas(hoja, columna_inicio)
    _proteger_hoja(hoja)

    return libro


def _escribir_encabezado(
    hoja, columnas: list, fila_encabezado: int, columna_inicio: int
) -> None:
    for indice, columna in enumerate(columnas):
        celda = hoja.cell(row=fila_encabezado, column=columna_inicio + indice)
        celda.value = columna["titulo"]
        aplicar_estilo_encabezado_tabla(celda)
        celda.protection = Protection(locked=True)


def _preparar_celdas_desbloqueadas(
    hoja,
    fila_encabezado: int,
    fila_inicio_datos: int,
    filas_datos: int,
    columna_inicio: int,
    columnas: list,
) -> None:
    """
    Desbloquea todas las celdas de datos para que la protección de hoja
    solo bloquee las celdas autonuméricas, que se re-bloquean luego.
    """
    proteccion_desbloqueada = Protection(locked=False)
    fila_fin_datos = fila_inicio_datos + filas_datos
    for numero_fila in range(fila_inicio_datos, fila_fin_datos):
        for indice_columna in range(len(columnas)):
            celda = hoja.cell(
                row=numero_fila, column=columna_inicio + indice_columna
            )
            celda.protection = proteccion_desbloqueada


def _configurar_columna(
    hoja,
    columna: dict,
    col_excel: int,
    fila_inicio_datos: int,
    filas_datos: int,
    reglamento: dict,
) -> None:
    tipo = columna["tipo"]

    if tipo == "autonumerico":
        _escribir_autonumerico(hoja, col_excel, fila_inicio_datos, filas_datos)
    elif tipo == "lista_desplegable":
        _aplicar_validacion_lista(
            hoja, columna, col_excel, fila_inicio_datos, filas_datos, reglamento
        )
    elif tipo == "entero_positivo":
        _aplicar_validacion_entero_positivo(
            hoja, columna, col_excel, fila_inicio_datos, filas_datos
        )

    if columna.get("fondo_gris_fijo"):
        _aplicar_fondo_gris(hoja, col_excel, fila_inicio_datos, filas_datos)


def _escribir_autonumerico(
    hoja, col_excel: int, fila_inicio_datos: int, filas_datos: int
) -> None:
    proteccion_bloqueada = Protection(locked=True)
    estilo_alineacion = Alignment(horizontal="center")

    fila_fin_datos = fila_inicio_datos + filas_datos
    for numero_fila in range(fila_inicio_datos, fila_fin_datos):
        numero_secuencial = numero_fila - fila_inicio_datos + 1
        celda = hoja.cell(row=numero_fila, column=col_excel)
        celda.value = numero_secuencial
        celda.protection = proteccion_bloqueada
        celda.alignment = estilo_alineacion


def _aplicar_validacion_lista(
    hoja,
    columna: dict,
    col_excel: int,
    fila_inicio_datos: int,
    filas_datos: int,
    reglamento: dict,
) -> None:
    opciones = _construir_opciones_lista(columna, reglamento)
    formula_opciones = '"' + ",".join(opciones) + '"'
    tipo_error = "stop" if columna.get("validacion_stop") else "warning"
    mensaje_error = columna.get("mensaje_error") or ""

    validacion = DataValidation(
        type="list",
        formula1=formula_opciones,
        allow_blank=True,
        showErrorMessage=bool(mensaje_error),
        error=mensaje_error,
        errorStyle=tipo_error,
    )
    hoja.add_data_validation(validacion)

    rango = _construir_referencia_rango_columna(
        col_excel, fila_inicio_datos, filas_datos
    )
    validacion.sqref = rango


def _construir_opciones_lista(columna: dict, reglamento: dict) -> list:
    if columna.get("fuente") == "yaml_load_types":
        return list(reglamento["load_types"].keys())
    return columna.get("opciones", [])


def _aplicar_validacion_entero_positivo(
    hoja,
    columna: dict,
    col_excel: int,
    fila_inicio_datos: int,
    filas_datos: int,
) -> None:
    mensaje_error = (
        columna.get("mensaje_error") or "Ingresá un número entero positivo."
    )

    validacion = DataValidation(
        type="whole",
        operator="greaterThan",
        formula1="0",
        allow_blank=True,
        showErrorMessage=True,
        error=mensaje_error,
        errorStyle="warning",
    )
    hoja.add_data_validation(validacion)

    rango = _construir_referencia_rango_columna(
        col_excel, fila_inicio_datos, filas_datos
    )
    validacion.sqref = rango


def _aplicar_fondo_gris(
    hoja, col_excel: int, fila_inicio_datos: int, filas_datos: int
) -> None:
    fila_fin_datos = fila_inicio_datos + filas_datos
    for numero_fila in range(fila_inicio_datos, fila_fin_datos):
        aplicar_estilo_celda_gris(hoja.cell(row=numero_fila, column=col_excel))


def _aplicar_borde_rango(
    hoja, fila_inicio: int, fila_fin: int, col_inicio: int, col_fin: int
) -> None:
    borde_fino = Side(style="thin")

    for numero_fila in range(fila_inicio, fila_fin + 1):
        for numero_col in range(col_inicio, col_fin + 1):
            celda = hoja.cell(row=numero_fila, column=numero_col)
            celda.border = Border(
                left=borde_fino,
                right=borde_fino,
                top=borde_fino,
                bottom=borde_fino,
            )


def _construir_referencia_rango_columna(
    col_excel: int, fila_inicio: int, cantidad_filas: int
) -> str:
    letra = get_column_letter(col_excel)
    return f"{letra}{fila_inicio}:{letra}{fila_inicio + cantidad_filas - 1}"


def _proteger_hoja(hoja) -> None:
    hoja.protection.sheet = True
    hoja.protection.selectLockedCells = False
    hoja.protection.selectUnlockedCells = False
    hoja.protection.formatColumns = False


def _escribir_panel_referencia(
    hoja,
    fila_encabezado: int,
    columna_inicio: int,
    columnas: list,
    reglamento: dict,
    configuracion_panel: dict,
) -> None:
    separador = configuracion_panel["columnas_separador"]
    columna_panel = columna_inicio + len(columnas) + separador
    fila_actual = fila_encabezado

    celda_titulo_permanentes = hoja.cell(
        row=fila_actual,
        column=columna_panel,
        value=configuracion_panel["titulo_permanentes"],
    )
    aplicar_estilo_encabezado_tabla(celda_titulo_permanentes)
    fila_actual += 1

    for nombre_tipo in reglamento["permanent_load_types"]:
        aplicar_estilo_label_metadata(
            hoja.cell(
                row=fila_actual,
                column=columna_panel,
                value=neutralizar_texto_libre(nombre_tipo),
            )
        )
        fila_actual += 1

    fila_actual += 1

    celda_titulo_tipos = hoja.cell(
        row=fila_actual,
        column=columna_panel,
        value=configuracion_panel["titulo_tipos_de_carga"],
    )
    aplicar_estilo_encabezado_tabla(celda_titulo_tipos)
    fila_actual += 1

    for id_tipo, datos_tipo in reglamento["load_types"].items():
        etiqueta = f"{id_tipo} - {datos_tipo['name']}"
        aplicar_estilo_label_metadata(
            hoja.cell(
                row=fila_actual,
                column=columna_panel,
                value=neutralizar_texto_libre(etiqueta),
            )
        )
        fila_actual += 1


def _guardar_archivo(libro: Workbook, ruta_destino: str) -> None:
    try:
        guardar_excel_atomico(libro, ruta_destino)
    except (PermissionError, OSError) as error:
        raise type(error)(
            f"No se pudo guardar la plantilla en '{ruta_destino}': {error}"
        ) from error
